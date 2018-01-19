# coding:utf-8
import sys
import platform
from openpyxl import load_workbook

# 先查看当前版本是否适合运行
curVer = platform.python_version();
if not(curVer.startswith("2.7")):
	print("当前 Python 版本是 " + curVer + "，暂不支持，请切换或安装 2.7.x 版本再重新运行")
	exit();

reload(sys)
sys.setdefaultencoding('utf-8')

input_file_path = raw_input("请输入文档的完整路径和文件名：")
start_row_column = raw_input("请输入第一个渠道号单元格的行数和列数(用\",\"隔开)：")
start_row = int(start_row_column.split(",")[0])
start_column = int(start_row_column.split(",")[1])

wb = load_workbook(filename=input_file_path.strip(), read_only=True)

sheets = wb.get_sheet_names()  # 获取所有表格(worksheet)的名字
ws = wb.get_sheet_by_name(sheets[0])  # 获取特定的 worksheet
print ("当前处理的表格是：" + sheets[0])

# 开始进行数据导出的准备
output_content = ""

i = start_row

while (i <= ws.max_row):
	A_column = str(ws.cell(row=i, column=start_column).value)
	print(A_column)
	i = i + 1
	output_content = output_content + A_column + "\n"

# 在当前目录输出
with open("./channels.txt", 'w') as f:
	f.write(output_content)
