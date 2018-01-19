# coding:utf-8
import sys
from openpyxl import load_workbook

reload(sys)
sys.setdefaultencoding('utf-8')

input_file_path = raw_input("请输入文档的完整路径和文件名：")
output_file_path = raw_input("请输入输出文档的保存路径：")

wb = load_workbook(filename=input_file_path.strip(), read_only=True)

sheets = wb.get_sheet_names()  # 获取所有表格(worksheet)的名字
ws = wb.get_sheet_by_name(sheets[2])  # 获取特定的 worksheet
print "当前处理的表格是：" + sheets[2]


# 表格数据检查，避免由于增删改导致数据导出异常
B_check = str(ws.cell(row=3, column=2).value)
if B_check != "模块/页面":
    print "B列有异常，请查看表格是否变更"
    exit()
C_check = str(ws.cell(row=3, column=3).value)
if C_check != "区域":
    print "C列有异常，请查看表格是否变更"
    exit()
D_check = str(ws.cell(row=3, column=4).value)
if D_check != "事件":
    print "D列有异常，请查看表格是否变更"
    exit()
L_check = str(ws.cell(row=3, column=12).value)
if L_check != "标识值eventKey":
    print "L列有异常，请查看表格是否变更"
    exit()

# 开始进行数据导出的准备
output_content = ""
B_before = ""
C_before = ""
D_before = ""

for i in range(5, ws.max_row):
    B_column = str(ws.cell(row=i, column=2).value)
    C_column = str(ws.cell(row=i, column=3).value)
    D_column = str(ws.cell(row=i, column=4).value)
    L_column = str(ws.cell(row=i, column=12).value)

    if L_column == "None":  # 如果标识值key(L列)==None，即后面已无数据，可跳出循环
        break

    if B_column != "None":
        B_before = B_column
    else:
        B_column = B_before

    if C_column != "None":
        C_before = C_column
    else:
        C_column = C_before

    if D_column != "None":
        D_before = D_column
    else:
        D_column = D_before

    one_line = L_column + "," + B_column + "_" + C_column + "_" + D_column + "," + "0" + "\n"  # 拼接

    print(one_line)
    output_content = output_content + one_line

with open(output_file_path + "/umeng.txt", 'w') as f:
    f.write(output_content)
